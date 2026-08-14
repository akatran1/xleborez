#!/usr/bin/env python3
"""Генератор каталога и карточек ножей для xleborez-new"""

import openpyxl
import json
import os
import re

XLSX = '/DATA/Progi/xleborez-new/Прайс ФФ Ножи рамные 20260709.xlsx'
OUT_DIR = '/DATA/Progi/xleborez-new/catalog/nozhi'
TEMPLATE_DETAIL = '/DATA/Progi/xleborez-new/catalog/nozhi/hrm-11/index.html'

# Load prices
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

def slugify(text):
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')

def short_name(name, brand):
    name = name.strip()
    # Extract the main dimensions
    match = re.match(r'[\d\s;×,\/\Ø]+\s*(.*)', name)
    if match and match.group(1).strip():
        short = match.group(1).strip()
    else:
        short = name[:60]
    if brand:
        return f"{brand}: {short}"
    return short

def parse_dimensions(name):
    """Extract dimensions from name string"""
    dims = {}
    # Length
    m = re.search(r'(\d+)\s*×', name)
    if m:
        dims['length'] = m.group(1)
    # Thickness
    m = re.search(r'×\s*([\d,]+)\s*×', name)
    if m:
        dims['thickness'] = m.group(1).replace(',', '.')
    # Width
    m = re.search(r'×\s*([\d,]+)\s*штифт', name)
    if m:
        dims['width'] = m.group(1).replace(',', '.')
    m = re.search(r'×\s*([\d,]+)\s*отверстие', name)
    if m:
        dims['width'] = m.group(1).replace(',', '.')
    # Pin
    m = re.search(r'штифт\s*([\d\s×]+)', name)
    if m:
        pin = m.group(1).strip()
        dims['pin'] = pin
    # Hole
    m = re.search(r'отверстие\s*[Ø]?\s*([\d\s]+)', name)
    if m:
        dims['pin'] = f'Ø {m.group(1).strip()} мм'
    return dims

# Collect products
steel_products = []
teflon_products = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[0] or '').strip()
    brand = str(row[1] or '').strip()
    price_steel = row[4]
    price_teflon = row[5]
    
    dims = parse_dimensions(name)
    
    if price_steel and name:
        slug = slugify(f"stal-{name[:40]}")
        steel_products.append({
            'id': slug,
            'name': name,
            'brand': brand,
            'price': int(price_steel),
            'coating': 'сталь',
            'dimensions': dims,
            'slug': slug,
        })
    
    if price_teflon and name:
        slug = slugify(f"teflon-{name[:40]}")
        teflon_products.append({
            'id': slug,
            'name': name,
            'brand': brand,
            'price': int(price_teflon),
            'coating': 'тефлон',
            'dimensions': dims,
            'slug': slug,
        })

all_products = steel_products + teflon_products

# Save catalog JSON
os.makedirs(f'{OUT_DIR}/assets', exist_ok=True)
with open(f'{OUT_DIR}/assets/catalog.json', 'w', encoding='utf-8') as f:
    json.dump(all_products, f, ensure_ascii=False, indent=2)

print(f"Собрано: {len(all_products)} позиций (сталь: {len(steel_products)}, тефлон: {len(teflon_products)})")

# =============================================
# GENERATE CATALOG PAGE
# =============================================

def make_card_html(p, coating_label):
    """Generate product card HTML"""
    img = f"/img/products/{p['id']}.jpg"
    # Check if image exists, use placeholder if not
    if not os.path.exists(f'/DATA/Progi/xleborez-new/img/products/{p["id"]}.jpg'):
        # Try to find any existing product image
        img = "/img/products/placeholder.svg"
    
    brand_html = f'<div class="product-card__brand">{p["brand"]}</div>' if p['brand'] else ''
    
    badge = 'Тефлон' if p['coating'] == 'тефлон' else 'Сталь'
    badge_class = ''
    
    return f'''        <div class="product-card" data-id="{p['id']}" data-price="{p['price']}">
          <a href="/catalog/nozhi/{p['slug']}/" class="product-card__image">
            <img src="{img}" alt="{p['name']}" loading="lazy">
            <span class="product-card__badge">{badge}</span>
          </a>
          <div class="product-card__body">
            {brand_html}
            <a href="/catalog/nozhi/{p['slug']}/" class="product-card__title">{p['name'][:55]}</a>
            <div class="product-card__price">{p['price']} ₽</div>
            <button class="product-card__btn add-to-cart"
              data-id="{p['id']}"
              data-name="{p['name'][:50]}"
              data-price="{p['price']}"
              data-image="{img}">В корзину</button>
          </div>
        </div>'''

steel_cards = '\n'.join(make_card_html(p, 'Сталь') for p in steel_products)
teflon_cards = '\n'.join(make_card_html(p, 'Тефлон') for p in teflon_products)

catalog_html = f'''<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Ножи для хлеборезки — каталог | ООО Фитфуд</title>
  <meta name="description" content="Каталог ножей для хлеборезок: стальные и тефлоновые. Цены от 158 рублей. Доставка по РФ.">
  <link rel="canonical" href="https://xleborez.ru/catalog/nozhi/">
  <link rel="icon" href="/favicon.ico">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Russo+One&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="/css/style.css">
  <style>
    .catalog-tabs {{ display: flex; gap: 0; margin-bottom: 2rem; border-bottom: 2px solid var(--color-gray-100); }}
    .catalog-tab {{ padding: 12px 28px; font-weight: 600; cursor: pointer; border-bottom: 3px solid transparent; transition: all var(--transition); color: var(--color-gray-500); }}
    .catalog-tab:hover {{ color: var(--color-dark); }}
    .catalog-tab.active {{ color: var(--color-dark); border-bottom-color: var(--color-gold); }}
    .catalog-section {{ display: none; }}
    .catalog-section.active {{ display: block; }}
    .catalog-count {{ font-size: 0.9rem; color: var(--color-gray-500); margin-bottom: 1.5rem; }}
  </style>
</head>
<body data-page="catalog">
  <header class="header">
    <div class="header__inner">
      <a href="/" class="header__logo"><img src="/img/logo-old.png" alt="ООО Фитфуд" style="height:50px;width:auto"></a>
      <nav class="header__nav">
        <a href="/">Главная</a>
        <a href="/catalog/nozhi/" class="active">Ножи</a>
        <a href="/catalog/zapchasti/">Запчасти</a>
        <a href="/catalog/hleborezki/">Хлеборезки</a>
        <a href="/catalog/klipsatory/">Клипсаторы</a>
        <a href="/contacts/">Контакты</a>
      </nav>
      <div class="header__actions">
        <a href="tel:89952226802" class="header__phone">8 995 222-68-02</a>
        <a href="/cart/" class="cart-btn">
          <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="9" cy="21" r="1"/><circle cx="20" cy="21" r="1"/><path d="M1 1h4l2.68 13.39a2 2 0 0 0 2 1.61h9.72a2 2 0 0 0 2-1.61L23 6H6"/></svg>
          <span class="cart-btn__count" id="cart-badge">0</span>
        </a>
        <div class="burger"><span></span><span></span><span></span></div>
      </div>
    </div>
  </header>

  <div class="container">
    <nav class="breadcrumbs">
      <a href="/">Главная</a>
      <span>›</span>
      <span>Ножи для хлеборезки</span>
    </nav>
  </div>

  <section class="section">
    <div class="container">
      <h1 style="margin-bottom:0.5rem;">Ножи для хлеборезки</h1>
      <p style="color:var(--color-gray-500);margin-bottom:2rem;">Цены с НДС. От 158 рублей. Доставка по всей России.</p>

      <!-- Tabs -->
      <div class="catalog-tabs">
        <div class="catalog-tab active" data-tab="steel">Стальные ({len(steel_products)})</div>
        <div class="catalog-tab" data-tab="teflon">Тефлоновые ({len(teflon_products)})</div>
      </div>

      <!-- Steel -->
      <div class="catalog-section active" id="tab-steel">
        <div class="catalog-count">Всего {len(steel_products)} позиций</div>
        <div class="product-grid">
{steel_cards}
        </div>
      </div>

      <!-- Teflon -->
      <div class="catalog-section" id="tab-teflon">
        <div class="catalog-count">Всего {len(teflon_products)} позиций</div>
        <div class="product-grid">
{teflon_cards}
        </div>
      </div>
    </div>
  </section>

  <section class="section section-alt">
    <div class="container">
      <div class="container-narrow" style="text-align:center;">
        <h2 style="margin-bottom:1rem;">Не нашли нужный нож?</h2>
        <p style="color:var(--color-gray-500);margin-bottom:1.5rem;">Скачайте чертёж, запишите размеры и пришлите на почту — мы подберём аналог.</p>
        <a href="https://yadi.sk/i/DsQXj-fI04jpag" target="_blank" class="btn btn--primary btn--lg">Скачать чертёж</a>
        <a href="mailto:info@xleborez.ru" class="btn btn--dark btn--lg" style="margin-left:0.5rem;">info@xleborez.ru</a>
      </div>
    </div>
  </section>

  <footer class="footer">
    <div class="container">
      <div class="footer__grid">
        <div class="footer__brand">
          <img src="/img/logo-old-2.png" alt="ООО Фитфуд" style="height:40px;width:auto">
          <p>Ножи для хлеборезок оптом и в розницу. Доставка по всей России.</p>
        </div>
        <div><h4>Каталог</h4><nav class="footer__nav">
          <a href="/catalog/nozhi/">Ножи для хлеборезки</a>
          <a href="/catalog/zapchasti/">Запчасти</a>
          <a href="/catalog/hleborezki/">Хлеборезки</a>
          <a href="/catalog/klipsatory/">Клипсаторы</a>
        </nav></div>
        <div><h4>Информация</h4><nav class="footer__nav">
          <a href="/contacts/">Контакты</a>
          <a href="/cart/">Корзина</a>
          <a href="/checkout/">Оформление заказа</a>
        </nav></div>
        <div><h4>Контакты</h4><nav class="footer__nav">
          <a href="tel:89952226802">8 995 222-68-02</a>
          <a href="mailto:info@xleborez.ru">info@xleborez.ru</a>
        </nav></div>
      </div>
      <div class="footer__bottom"><p>© 2015—<script>document.write(new Date().getFullYear())</script> ООО Фитфуд</p></div>
    </div>
  </footer>

  <script>
  document.addEventListener('DOMContentLoaded', function() {{
    const tabs = document.querySelectorAll('.catalog-tab');
    tabs.forEach(tab => {{
      tab.addEventListener('click', function() {{
        tabs.forEach(t => t.classList.remove('active'));
        this.classList.add('active');
        document.querySelectorAll('.catalog-section').forEach(s => s.classList.remove('active'));
        document.getElementById('tab-' + this.dataset.tab).classList.add('active');
      }});
    }});
  }});
  </script>
  <script src="/js/app.js"></script>

  <!-- Messenger Widget -->
  <script src="/vidget/messenger-widget.js"></script>
  <script>
    MessengerWidget.init({
      telegram: { username: 'andresemykin', link: 'https://t.me/andresemykin' },
      max: { username: 'ООО Фитфуд', link: 'https://max.ru/u/f9LHodD0cOKSzHGEQIX3HAhXGwgQALr4CxadoVlHiQKQ9ug17KGfP-cbEMU' }
    });
  </script>
</body>
</html>'''

with open(f'{OUT_DIR}/index.html', 'w', encoding='utf-8') as f:
    f.write(catalog_html)

print(f"Каталог: {OUT_DIR}/index.html — {len(steel_products)} стальных + {len(teflon_products)} тефлоновых")

# =============================================
# GENERATE PRODUCT PAGES
# =============================================

# Read template
with open(TEMPLATE_DETAIL, 'r', encoding='utf-8') as f:
    template = f.read()

def generate_product_page(p, all_products):
    """Generate product detail page from template"""
    dims = p['dimensions']
    coating_ru = 'тефлоновым покрытием' if p['coating'] == 'тефлон' else 'стальные'
    
    title = f"Ножи для хлеборезки {p['name'][:50]} — купить | ООО Фитфуд"
    desc = f"Ножи для хлеборезки {p['name'][:60]} — {p['price']} ₽. {coating_ru.capitalize()}. "
    if p['brand']:
        desc += f"Производитель: {p['brand']}. "
    desc += f"Доставка по РФ."
    
    # Dimensions table
    dims_html = ''
    if dims.get('length'):
        dims_html += f'<li>Длина ножа: {dims["length"]} мм</li>'
    if dims.get('thickness'):
        dims_html += f'<li>Толщина полотна: {dims["thickness"]} мм</li>'
    if dims.get('width'):
        dims_html += f'<li>Ширина полотна: {dims["width"]} мм</li>'
    if dims.get('pin'):
        dims_html += f'<li>Крепление: {dims["pin"]}</li>'
    
    dims_html += f'<li>Покрытие: {coating_ru}</li>'
    
    brand_display = p['brand'] if p['brand'] else 'Универсальные'
    
    # Similar products (4 from same coating type)
    similar = [pr for pr in all_products if pr['coating'] == p['coating'] and pr['id'] != p['id']][:4]
    similar_html = ''
    for pr in similar:
        img = f"/img/products/{pr['id']}.jpg"
        if not os.path.exists(f'/DATA/Progi/xleborez-new/img/products/{pr["id"]}.jpg'):
            img = "/img/products/placeholder.svg"
        brand_html = f'<div class="product-card__brand">{pr["brand"]}</div>' if pr['brand'] else ''
        similar_html += f'''
        <div class="product-card" data-id="{pr['id']}" data-price="{pr['price']}">
          <a href="/catalog/nozhi/{pr['slug']}/" class="product-card__image">
            <img src="{img}" alt="{pr['name'][:40]}" loading="lazy">
            <span class="product-card__badge">{"Тефлон" if pr['coating'] == 'тефлон' else "Сталь"}</span>
          </a>
          <div class="product-card__body">
            {brand_html}
            <a href="/catalog/nozhi/{pr['slug']}/" class="product-card__title">{pr['name'][:50]}</a>
            <div class="product-card__price">{pr['price']} ₽</div>
            <button class="product-card__btn add-to-cart"
              data-id="{pr['id']}" data-name="{pr['name'][:50]}"
              data-price="{pr['price']}" data-image="{img}">В корзину</button>
          </div>
        </div>'''
    
    # Image
    img_path = f"/img/products/{p['id']}.jpg"
    if not os.path.exists(f'/DATA/Progi/xleborez-new/img/products/{p["id"]}.jpg'):
        img_path = "/img/products/placeholder.svg"
    
    page = f'''<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title}</title>
  <meta name="description" content="{desc}">
  <link rel="canonical" href="https://xleborez.ru/catalog/nozhi/{p['slug']}/">
  <script type="application/ld+json">{{
    "@context": "https://schema.org",
    "@type": "Product",
    "name": "{p['name'][:60]}",
    "image": "https://xleborez.ru{img_path}",
    "brand": {{"@type": "Brand", "name": "{brand_display}"}},
    "offers": {{"@type": "Offer", "price": "{p['price']}", "priceCurrency": "RUB", "availability": "https://schema.org/InStock", "url": "https://xleborez.ru/catalog/nozhi/{p['slug']}/"}}
  }}</script>
  <link rel="icon" href="/favicon.ico">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Russo+One&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="/css/style.css">
</head>
<body>
  <header class="header">
    <div class="header__inner">
      <a href="/" class="header__logo"><img src="/img/logo-old.png" alt="ООО Фитфуд" style="height:50px;width:auto"></a>
      <nav class="header__nav">
        <a href="/">Главная</a>
        <a href="/catalog/nozhi/" class="active">Ножи</a>
        <a href="/catalog/zapchasti/">Запчасти</a>
        <a href="/catalog/hleborezki/">Хлеборезки</a>
        <a href="/catalog/klipsatory/">Клипсаторы</a>
        <a href="/contacts/">Контакты</a>
      </nav>
      <div class="header__actions">
        <a href="tel:89952226802" class="header__phone">8 995 222-68-02</a>
        <a href="/cart/" class="cart-btn">
          <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="9" cy="21" r="1"/><circle cx="20" cy="21" r="1"/><path d="M1 1h4l2.68 13.39a2 2 0 0 0 2 1.61h9.72a2 2 0 0 0 2-1.61L23 6H6"/></svg>
          <span class="cart-btn__count" id="cart-badge">0</span>
        </a>
        <div class="burger"><span></span><span></span><span></span></div>
      </div>
    </div>
  </header>

  <div class="container">
    <nav class="breadcrumbs">
      <a href="/">Главная</a>
      <span>›</span>
      <a href="/catalog/nozhi/">Ножи для хлеборезки</a>
      <span>›</span>
      <span>{p['name'][:55]}</span>
    </nav>
  </div>

  <section class="product-detail">
    <div class="container" style="max-width:var(--container-max);margin:0 auto;padding:0 var(--space-lg);display:grid;grid-template-columns:1fr 1fr;gap:3rem;">
      <div>
        <div class="product-detail__gallery">
          <div class="product-detail__main-image">
            <img id="main-product-image" src="{img_path}" alt="{p['name'][:50]}">
          </div>
        </div>
      </div>
      <div class="product-detail__info">
        <h1>Ножи для хлеборезки {p['name'][:60]}</h1>
        <div class="product-detail__price">{p['price']} ₽</div>
        <div class="product-detail__meta">
          <div class="product-detail__meta-item">
            <span class="product-detail__meta-label">Бренд</span>
            <span class="product-detail__meta-value">{brand_display}</span>
          </div>
          <div class="product-detail__meta-item">
            <span class="product-detail__meta-label">Наличие</span>
            <span class="product-detail__meta-value" style="color:var(--color-success);">В наличии</span>
          </div>
          <div class="product-detail__meta-item">
            <span class="product-detail__meta-label">Покрытие</span>
            <span class="product-detail__meta-value">{coating_ru.capitalize()}</span>
          </div>
        </div>

        <div class="product-detail__desc">
          <p><strong>Характеристики:</strong></p>
          <ul style="list-style:disc;padding-left:1.5rem;line-height:2;">
            {dims_html}
          </ul>
          <p style="margin-top:1rem;">
            Ножи из высококачественной стали. Цена указана с НДС.
            Подходит для хлеборезок соответствующих моделей.
          </p>
          <p style="margin-top:1rem;">
            <a href="https://yadi.sk/i/DsQXj-fI04jpag" target="_blank" style="color:var(--color-gold);font-weight:600;">
              Скачать чертёж ножа →
            </a>
          </p>
        </div>

        <div class="product-detail__actions">
          <div class="quantity-selector">
            <button class="qty-minus">−</button>
            <input type="number" value="1" min="1" max="999">
            <button class="qty-plus">+</button>
          </div>
          <button class="btn btn--primary btn--lg add-to-cart"
            data-id="{p['id']}"
            data-name="{p['name'][:50]}"
            data-price="{p['price']}"
            data-image="{img_path}">
            В корзину
          </button>
        </div>
      </div>
    </div>
  </section>

  <!-- Similar -->
  <section class="section section-alt">
    <div class="container">
      <div class="section-title"><h2>Похожие товары</h2></div>
      <div class="product-grid">
{similar_html}
      </div>
    </div>
  </section>

  <footer class="footer">
    <div class="container">
      <div class="footer__grid">
        <div class="footer__brand">
          <img src="/img/logo-old-2.png" alt="ООО Фитфуд" style="height:40px;width:auto">
          <p>Ножи для хлеборезок оптом и в розницу. Доставка по всей России.</p>
        </div>
        <div><h4>Каталог</h4><nav class="footer__nav">
          <a href="/catalog/nozhi/">Ножи для хлеборезки</a>
          <a href="/catalog/zapchasti/">Запчасти</a>
          <a href="/catalog/hleborezki/">Хлеборезки</a>
          <a href="/catalog/klipsatory/">Клипсаторы</a>
        </nav></div>
        <div><h4>Информация</h4><nav class="footer__nav">
          <a href="/contacts/">Контакты</a>
          <a href="/cart/">Корзина</a>
          <a href="/checkout/">Оформление заказа</a>
        </nav></div>
        <div><h4>Контакты</h4><nav class="footer__nav">
          <a href="tel:89952226802">8 995 222-68-02</a>
          <a href="mailto:info@xleborez.ru">info@xleborez.ru</a>
        </nav></div>
      </div>
      <div class="footer__bottom"><p>© 2015—<script>document.write(new Date().getFullYear())</script> ООО Фитфуд</p></div>
    </div>
  </footer>

  <script src="/js/app.js"></script>

  <!-- Messenger Widget -->
  <script src="/vidget/messenger-widget.js"></script>
  <script>
    MessengerWidget.init({
      telegram: { username: 'andresemykin', link: 'https://t.me/andresemykin' },
      max: { username: 'ООО Фитфуд', link: 'https://max.ru/u/f9LHodD0cOKSzHGEQIX3HAhXGwgQALr4CxadoVlHiQKQ9ug17KGfP-cbEMU' }
    });
  </script>
</body>
</html>'''

    return page

# Generate all product pages
generated = 0
for p in all_products:
    prod_dir = f'{OUT_DIR}/{p["slug"]}'
    os.makedirs(prod_dir, exist_ok=True)
    
    html = generate_product_page(p, all_products)
    with open(f'{prod_dir}/index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    generated += 1
    if generated <= 3:
        print(f"  {generated}. {p['slug']}/ — {p['name'][:50]} — {p['price']} ₽")

print(f"\nСгенерировано страниц товаров: {generated}")

# Update sitemap
sitemap_entries = []
for p in all_products:
    sitemap_entries.append(f'''  <url>
    <loc>https://xleborez.ru/catalog/nozhi/{p['slug']}/</loc>
    <changefreq>monthly</changefreq>
    <priority>0.8</priority>
  </url>''')

sitemap = f'''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url>
    <loc>https://xleborez.ru/</loc>
    <changefreq>weekly</changefreq>
    <priority>1.0</priority>
  </url>
  <url>
    <loc>https://xleborez.ru/catalog/nozhi/</loc>
    <changefreq>weekly</changefreq>
    <priority>0.9</priority>
  </url>
{chr(10).join(sitemap_entries)}
  <url>
    <loc>https://xleborez.ru/catalog/zapchasti/</loc>
    <changefreq>weekly</changefreq>
    <priority>0.8</priority>
  </url>
  <url>
    <loc>https://xleborez.ru/catalog/hleborezki/</loc>
    <changefreq>weekly</changefreq>
    <priority>0.8</priority>
  </url>
  <url>
    <loc>https://xleborez.ru/catalog/klipsatory/</loc>
    <changefreq>weekly</changefreq>
    <priority>0.8</priority>
  </url>
  <url>
    <loc>https://xleborez.ru/contacts/</loc>
    <changefreq>monthly</changefreq>
    <priority>0.6</priority>
  </url>
</urlset>'''

with open('/DATA/Progi/xleborez-new/sitemap.xml', 'w', encoding='utf-8') as f:
    f.write(sitemap)
print(f"Sitemap обновлён: {len(sitemap_entries)} позиций")
